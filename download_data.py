from openpyxl import load_workbook
from setting import SETTINGS
import os
import requests
import zipfile

forest_dir = SETTINGS.forest_data_dir
# Load the CSV data
forest_data = os.path.join(forest_dir, 'Initial_forest_data.xlsx')


def get_hyperlinks():
    workbook = load_workbook(forest_data, data_only=True)
    # Select the first sheet
    sheet = workbook.active

    for row in sheet.iter_rows(min_col=1, max_col=1):
        for cell in row:
            if cell.hyperlink:  # Check if the cell contains a hyperlink
                print('yield, ', cell.value, cell.hyperlink.target)
                yield cell.value, cell.hyperlink.target


def download_unzip(file_name, file_url):
    response = requests.get(file_url)
    file_name = file_name.replace('/', '.')
    zip_filepath = os.path.join(forest_dir, f'{file_name}.zip')
    print('zip_filepath', zip_filepath)
    if os.path.isfile(zip_filepath):
        print('zip exists')
        return
    with open(zip_filepath, 'wb') as f:
        f.write(response.content)

    target_unzip_folder = os.path.join(forest_dir, file_name)
    with zipfile.ZipFile(zip_filepath, 'r') as zip_ref:
        zip_ref.extractall(target_unzip_folder)


def main():
    for name, hyperlink in get_hyperlinks():
        download_unzip(name, hyperlink)

main()